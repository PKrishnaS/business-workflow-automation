# ============================================================
# report_gen/excel_report.py — Formatted Excel report generation
# ============================================================
# Creates styled .xlsx files with multiple sheets:
#   Sheet 1: Summary stats
#   Sheet 2: Cleaned data
#   Sheet 3: Group-by analysis
# ============================================================

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Union

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from config.settings import (
    PDF_PRIMARY_COLOR, PDF_SECONDARY_COLOR,
    DATA_OUTPUT_DIR, REPORT_COMPANY_NAME
)
from utils.logger import get_logger, log_function_call
from utils.helpers import get_timestamp, sanitize_filename

logger = get_logger(__name__)


def _hex_from_rgb(r, g, b):
    """Convert RGB tuple to Excel's hex color format (no #)."""
    return f"{r:02X}{g:02X}{b:02X}"


# Color constants for styling
PRIMARY_HEX   = _hex_from_rgb(*PDF_PRIMARY_COLOR)
SECONDARY_HEX = _hex_from_rgb(*PDF_SECONDARY_COLOR)
LIGHT_BLUE    = "EBF3FF"
LIGHT_GRAY    = "F5F5F5"
MID_GRAY      = "CCCCCC"
WHITE         = "FFFFFF"
DARK_TEXT     = "2B2B2B"


class ExcelReport:
    """
    Create a professional multi-sheet Excel report.

    HOW TO USE:
        report = ExcelReport(title="Q1 Sales Report")
        report.add_summary_sheet(stats_dict)
        report.add_data_sheet("Sales Data", sales_df)
        path = report.save()
    """

    def __init__(self, title: str, output_dir: Union[str, Path] = None):
        self.title      = title
        self.output_dir = Path(output_dir) if output_dir else DATA_OUTPUT_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.wb         = Workbook()
        self.wb.remove(self.wb.active)   # Remove the default empty sheet

    def _make_border(self, style="thin"):
        """Create a cell border."""
        side = Side(style=style, color=MID_GRAY)
        return Border(left=side, right=side, top=side, bottom=side)

    def _make_header_fill(self):
        """Primary color background for header rows."""
        return PatternFill("solid", fgColor=PRIMARY_HEX)

    def _make_alt_fill(self):
        """Light blue alternating row fill."""
        return PatternFill("solid", fgColor=LIGHT_BLUE)

    def _make_gray_fill(self):
        return PatternFill("solid", fgColor=LIGHT_GRAY)

    def _style_header_cell(self, cell, text: str, font_size: int = 11):
        """Apply header styling to a cell."""
        cell.value          = text
        cell.font           = Font(bold=True, color=WHITE, size=font_size, name="Calibri")
        cell.fill           = self._make_header_fill()
        cell.alignment      = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border         = self._make_border()

    def _style_data_cell(self, cell, alternate: bool = False):
        """Apply data row styling to a cell."""
        cell.fill      = self._make_alt_fill() if alternate else PatternFill("solid", fgColor=WHITE)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border    = self._make_border()
        cell.font      = Font(size=10, name="Calibri", color=DARK_TEXT)

    def _auto_fit_columns(self, ws, min_width: int = 10, max_width: int = 40):
        """Automatically set column widths based on content."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)

    def add_summary_sheet(self, metrics: dict, sheet_name: str = "Summary") -> "ExcelReport":
        """
        Add a summary sheet with key-value metric pairs.

        Args:
            metrics:    dict of {label: value} pairs
            sheet_name: Name for this sheet tab
        """
        ws = self.wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False

        # ── Title block ──────────────────────────────────────
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value     = self.title
        title_cell.font      = Font(bold=True, size=16, color=PRIMARY_HEX, name="Calibri")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 35

        # Generated timestamp
        ws["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y at %H:%M')} | {REPORT_COMPANY_NAME}"
        ws["A2"].font      = Font(size=9, color=SECONDARY_HEX, italic=True, name="Calibri")
        ws["A2"].alignment = Alignment(horizontal="left")

        # Separator row
        ws.row_dimensions[3].height = 8

        # ── Column headers ───────────────────────────────────
        headers = ["Metric", "Value", "Metric", "Value"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx)
            self._style_header_cell(cell, header)

        ws.row_dimensions[4].height = 22

        # ── Metric rows (2 per row for compactness) ───────────
        keys   = list(metrics.keys())
        values = list(metrics.values())
        row_num = 5

        for i in range(0, len(keys), 2):
            is_alt = (i // 2) % 2 == 0

            # Left pair
            label_cell = ws.cell(row=row_num, column=1, value=str(keys[i]))
            value_cell = ws.cell(row=row_num, column=2, value=values[i])
            label_cell.font   = Font(bold=True, size=10, name="Calibri", color=DARK_TEXT)
            label_cell.fill   = self._make_gray_fill() if is_alt else PatternFill("solid", fgColor=WHITE)
            value_cell.fill   = self._make_gray_fill() if is_alt else PatternFill("solid", fgColor=WHITE)
            value_cell.font   = Font(size=10, name="Calibri", color=PRIMARY_HEX, bold=True)
            label_cell.border = value_cell.border = self._make_border()
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            value_cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

            # Right pair (if exists)
            if i + 1 < len(keys):
                label_cell2 = ws.cell(row=row_num, column=3, value=str(keys[i+1]))
                value_cell2 = ws.cell(row=row_num, column=4, value=values[i+1])
                label_cell2.font   = Font(bold=True, size=10, name="Calibri", color=DARK_TEXT)
                label_cell2.fill   = self._make_gray_fill() if is_alt else PatternFill("solid", fgColor=WHITE)
                value_cell2.fill   = self._make_gray_fill() if is_alt else PatternFill("solid", fgColor=WHITE)
                value_cell2.font   = Font(size=10, name="Calibri", color=PRIMARY_HEX, bold=True)
                label_cell2.border = value_cell2.border = self._make_border()
                label_cell2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                value_cell2.alignment = Alignment(horizontal="right", vertical="center", indent=1)

            ws.row_dimensions[row_num].height = 20
            row_num += 1

        self._auto_fit_columns(ws)
        return self

    def add_data_sheet(self, sheet_name: str, df: pd.DataFrame,
                       max_rows: int = 10000) -> "ExcelReport":
        """
        Add a formatted data table sheet from a pandas DataFrame.

        Args:
            sheet_name: Name for this sheet tab.
            df:         pandas DataFrame to display.
            max_rows:   Maximum rows to include (default 10,000).
        """
        ws = self.wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False

        if df.empty:
            ws["A1"] = "No data available."
            return self

        display_df = df.head(max_rows)

        # ── Header row ────────────────────────────────────────
        headers = [str(col).replace("_", " ").title() for col in display_df.columns]
        for col_idx, header in enumerate(headers, start=1):
            self._style_header_cell(ws.cell(row=1, column=col_idx), header)

        ws.row_dimensions[1].height = 24

        # ── Data rows ─────────────────────────────────────────
        for row_idx, (_, row) in enumerate(display_df.iterrows(), start=2):
            is_alt = (row_idx % 2 == 0)
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                self._style_data_cell(cell, alternate=is_alt)
                # Right-align numbers
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            ws.row_dimensions[row_idx].height = 18

        # Freeze the header row so it stays visible while scrolling
        ws.freeze_panes = ws["A2"]

        # Auto-filter dropdown on header row
        ws.auto_filter.ref = ws.dimensions

        self._auto_fit_columns(ws)
        return self

    @log_function_call(logger)
    def save(self, filename: str = None) -> Path:
        """
        Save the workbook to a .xlsx file.

        Args:
            filename: Custom filename without extension. Auto-generated if None.

        Returns:
            Path to the saved file.
        """
        if not self.wb.sheetnames:
            raise RuntimeError("No sheets added. Call add_summary_sheet() or add_data_sheet() first.")

        if filename is None:
            filename = f"{sanitize_filename(self.title)}_{get_timestamp()}"
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        output_path = self.output_dir / filename
        self.wb.save(str(output_path))
        logger.info(f"Excel report saved: {output_path}")
        return output_path
